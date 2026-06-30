import os
import re
import time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TARGET_COLUMNS = [
    "Юрлицо / компания",
    "Роль",
    "Бренд(ы) в категории",
    "Подкатегории / товары из категории",
    "Что производит / функция",
    "УНП",
    "Источник",
]


def clean_text(value):
    if value is None:
        return ""

    value = str(value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def safe_sheet_name(name):
    name = re.sub(r"[\[\]\:\*\?\/\\]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:31] or "Поставщики"


def autosize_columns(ws, max_width=55):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for cell in ws[col_letter]:
            value = clean_text(cell.value)
            max_len = max(max_len, len(value))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), max_width)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    subtitle_fill = PatternFill("solid", fgColor="EAF3F8")

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(TARGET_COLUMNS),
    )

    ws["A2"].font = Font(italic=True, color="666666")
    ws["A2"].fill = subtitle_fill
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=len(TARGET_COLUMNS),
    )

    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(
        min_row=4,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"

    autosize_columns(ws)

    for row_idx in range(4, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 48


def create_excel_file(rows, output_path, source_url):
    output_dir = os.path.dirname(output_path)

    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name("Поставщики")

    ws.append(["Поставщики, производители и бренды"])
    ws.append([
        f"Источник: {source_url} · "
        f"Сформировано автоматически: {time.strftime('%Y-%m-%d %H:%M:%S')}"
    ])
    ws.append(TARGET_COLUMNS)

    for row in rows:
        ws.append([row.get(col, "") for col in TARGET_COLUMNS])

    style_sheet(ws)

    info = wb.create_sheet("README")
    info.append(["Как читать файл"])
    info.append([
        "1",
        "Если указано «Не найдено на карточке товара», значит сайт не отдал производителя/поставщика в доступных данных.",
    ])
    info.append([
        "2",
        "Для юридически точного поставщика нужны карточка товара, этикетка, договорные данные сети или данные маркировки.",
    ])
    info.append([
        "3",
        "УНП автоматически не заполняется, потому что не каждый сайт отдаёт юридические реквизиты поставщика.",
    ])
    info.append([
        "4",
        "Колонка «Источник» содержит страницу или API-источник, откуда был найден товар.",
    ])
    info.append([
        "5",
        "Универсальный парсер работает лучше всего со страницами каталогов, категорий и поиска.",
    ])

    autosize_columns(info)

    wb.save(output_path)

    return output_path
